import joblib
import pandas as pd
import numpy as np
import os
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List
from sqlalchemy.orm import Session
from fastapi import Depends
import database
import sql_models
from database import engine, get_db


# 1. Initialize API
app = FastAPI(title="Bearing Failure Prediction API", version="1.0")

# Enable CORS (So React can talk to Python)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create the tables automatically if they don't exist
sql_models.Base.metadata.create_all(bind=engine)

# Allow the React frontend to talk to this API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, replace with your React app URL
    allow_methods=["*"],
    allow_headers=["*"],
)

model_filename = 'universal_model.joblib'

# Check if it's in a 'models' folder just in case
if not os.path.exists(model_filename):
    # Try looking in a models subdirectory
    model_filename = os.path.join('models', 'universal_model.joblib')

# 2. Debugging Print - This will tell us where it is looking
print(f"🔍 Looking for model at: {os.path.abspath(model_filename)}")

if os.path.exists(model_filename):
    print("✅ Model found! Loading...")
    try:
        package = joblib.load(model_filename)
        model = package['model']
        scaler = package['scaler']
        threshold = package['threshold']
    except Exception as e:
        raise RuntimeError(f"File found but failed to load: {e}")
else:
    # 3. CRITICAL ERROR MESSAGE
    current_dir = os.getcwd()
    files_here = os.listdir(current_dir)
    raise RuntimeError(f"\n❌ CRITICAL ERROR: Model file not found.\n   -> I looked for: {model_filename}\n   -> Current Folder: {current_dir}\n   -> Files in this folder: {files_here}")

# --- Load simulation data ---
CURRENT_INDEX = 0
DATA_STREAM = []
TOTAL_RECORDS = 0

# Try to load full dataset first, fall back to synthetic
data_file = "simulation_data.csv"
if not os.path.exists(data_file):
    data_file = "synthetic_test_data.csv"

if os.path.exists(data_file):
    print(f"📊 Loading simulation data from: {data_file}")
    df = pd.read_csv(data_file)
    DATA_STREAM = df.to_dict('records')
    TOTAL_RECORDS = len(DATA_STREAM)
    print(f"✅ Loaded {TOTAL_RECORDS} records for simulation")
else:
    print(f"⚠️ Warning: No simulation data found. Run train.py first.")


# 3. Define Input Schema (for single sensor reading simulation)
class SensorReading(BaseModel):
    # Expecting a list of 20 floats
    features: List[float] 

# Alias for the predict endpoint
PredictionInput = SensorReading 

# --- ENDPOINT 1: Health Check ---
@app.get("/")
def home():
    return {"status": "System Online", "model": "Universal PCA v1.0"}

# --- NEW: Get Model Info (threshold) ---
@app.get("/model-info")
def get_model_info():
    return {
        "threshold": float(threshold),
        "model_type": "PCA",
        "features": 20
    }

@app.post("/predict")
def predict(input_data: PredictionInput, db: Session = Depends(get_db)): # <--- Added db dependency
    try:
        # 1. Run the Math (Same as before)
        features = np.array(input_data.features).reshape(1, -1)
        features_scaled = scaler.transform(features)
        recon = model.inverse_transform(model.transform(features_scaled))
        mse = np.mean(np.square(features_scaled - recon), axis=1)[0]
        
        # 2. Determine Status
        status = "CRITICAL_FAILURE" if mse > threshold else "HEALTHY"
        
        # 3. --- NEW: SAVE TO DATABASE ---
        # Create a new record object
        new_record = sql_models.PredictionRecord(
            input_features=input_data.features, # Save the input numbers
            anomaly_score=float(mse),           # Save the score
            threshold=float(threshold),         # Save the limit
            status=status                       # Save the result
        )
        
        # Add to the "Session" and "Commit" (Save)
        db.add(new_record)
        db.commit()
        db.refresh(new_record) # Get the ID back
        
        # 4. Return Response (Same as before)
        return {
            "id": new_record.id, # We can now return the Database ID!
            "anomaly_score": float(mse),
            "threshold": float(threshold),
            "status": status,
            "alert": bool(mse > threshold)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ENDPOINT 3: Batch Analysis (File Upload) ---
@app.post("/analyze-batch")
async def analyze_batch(file: UploadFile = File(...)):
    # Read uploaded CSV
    try:
        df = pd.read_csv(file.file)
        
        # Validate columns (Simplistic check)
        if df.shape[1] < 20:
             raise HTTPException(status_code=400, detail="CSV must have at least 20 feature columns")
        
        # Select numeric features
        features = df.select_dtypes(include=[np.number]).iloc[:, :20]
        
        # Process
        X_scaled = scaler.transform(features)
        X_recon = model.inverse_transform(model.transform(X_scaled))
        mse_scores = np.mean(np.square(X_scaled - X_recon), axis=1)
        
        # Results
        results = []
        for i, score in enumerate(mse_scores):
            results.append({
                "index": i,
                "score": round(score, 4),
                "is_anomaly": bool(score > threshold)
            })
            
        return {"filename": file.filename, "total_records": len(df), "results": results}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ENDPOINT 4: Get All Records ---
@app.get("/records")
def get_records(db: Session = Depends(get_db)):
    records = db.query(sql_models.PredictionRecord).all()
    return {"records": records}

# --- ENDPOINT 5: Get History ---
@app.get("/history")
def get_history(db: Session = Depends(get_db)):
    # SQL Query: "SELECT * FROM predictions ORDER BY id DESC LIMIT 50"
    records = db.query(sql_models.PredictionRecord).order_by(sql_models.PredictionRecord.id.desc()).limit(50).all()
    return records

# --- ENDPOINT 6: Health Check ---
@app.get("/health-check")
def health_check(db: Session = Depends(get_db)):
    """Diagnose system health: check database, model, and predictions table"""
    health_status = {
        "database": "UNKNOWN",
        "model": "UNKNOWN", 
        "predictions_table": "UNKNOWN",
        "total_predictions": 0,
        "last_predictions": [],
        "simulation_data": "UNKNOWN",
        "simulation_records": 0
    }
    
    # Check model
    try:
        health_status["model"] = "OK" if model is not None else "NOT_LOADED"
    except:
        health_status["model"] = "ERROR"
    
    # Check simulation data
    try:
        health_status["simulation_records"] = TOTAL_RECORDS
        health_status["simulation_data"] = "OK" if TOTAL_RECORDS > 0 else "EMPTY"
    except:
        health_status["simulation_data"] = "ERROR"
    
    # Check database and predictions table
    try:
        # Try to count records
        count = db.query(sql_models.PredictionRecord).count()
        health_status["total_predictions"] = count
        health_status["database"] = "OK"
        health_status["predictions_table"] = "OK" if count >= 0 else "ERROR"
        
        # Get last 5 predictions
        last_records = db.query(sql_models.PredictionRecord).order_by(
            sql_models.PredictionRecord.id.desc()
        ).limit(5).all()
        
        health_status["last_predictions"] = [
            {
                "id": r.id,
                "status": r.status,
                "score": r.anomaly_score,
                "threshold": r.threshold
            } for r in last_records
        ]
        
    except Exception as e:
        health_status["database"] = f"ERROR: {str(e)}"
        health_status["predictions_table"] = "ERROR"
    
    # Overall status
    all_ok = (
        health_status["database"] == "OK" and 
        health_status["model"] == "OK" and 
        health_status["predictions_table"] == "OK"
    )
    health_status["overall"] = "HEALTHY" if all_ok else "ISSUES_DETECTED"
    
    return health_status


@app.post("/simulate/reset")
def reset_simulation():
    """Rewind the 'movie' to the beginning"""
    global CURRENT_INDEX
    CURRENT_INDEX = 0
    return {"message": "Simulation reset to Day 1", "total_records": len(DATA_STREAM)}

@app.get("/simulate/next")
def get_next_real_reading(db: Session = Depends(get_db)):
    """Get the NEXT row of real data from the NASA experiment"""
    global CURRENT_INDEX
    
    if not DATA_STREAM:
        raise HTTPException(status_code=404, detail="No data loaded")
    
    if CURRENT_INDEX >= len(DATA_STREAM):
        return {"finished": True, "total_records": TOTAL_RECORDS}

    # 1. Get the row
    row = DATA_STREAM[CURRENT_INDEX]
    CURRENT_INDEX += 1

    # 2. Extract features (exclude timestamp/dataset columns)
    # We define the 20 columns we need (based on features.py logic)
    # Note: features.py outputs specific column names, we just grab values 0-19
    # excluding 'timestamp' and 'dataset'
    feature_values = [v for k, v in row.items() if k not in ['timestamp', 'dataset']]
    
    # Ensure we strictly have 20 features (sometimes ingestion output varies)
    # We assume the first 20 numeric columns are the features.
    feature_values = feature_values[:20] 

    # 3. Run Prediction logic (Same as /predict)
    features_arr = np.array(feature_values).reshape(1, -1)
    features_scaled = scaler.transform(features_arr)
    recon = model.inverse_transform(model.transform(features_scaled))
    mse = np.mean(np.square(features_scaled - recon), axis=1)[0]
    
    thresh_val = float(threshold)
    mse_val = float(mse)
    status = "CRITICAL_FAILURE" if mse_val > thresh_val else "HEALTHY"

    # 4. Save to DB
    new_record = sql_models.PredictionRecord(
        input_features=feature_values,
        anomaly_score=mse_val,
        threshold=thresh_val,
        status=status
    )
    db.add(new_record)
    db.commit()
    db.refresh(new_record)

    return {
        "finished": False,
        "index": CURRENT_INDEX,
        "total_records": TOTAL_RECORDS,
        "timestamp": str(row['timestamp']),
        "anomaly_score": mse_val,
        "threshold": thresh_val,
        "status": status,
        "alert": bool(mse_val > thresh_val)
    }

# ============================================================
# CUSTOM DATA UPLOAD & SIMULATION ENDPOINTS
# ============================================================

# Import data transformer
from data_transformer import DataTransformer, transform_data

# Global state for custom uploaded data simulation
CUSTOM_DATA_STREAM = []
CUSTOM_CURRENT_INDEX = 0
CUSTOM_TOTAL_RECORDS = 0
CUSTOM_TRANSFORMATION_INFO = {}

@app.post("/preprocess-upload")
async def preprocess_upload(file: UploadFile = File(...), window_size: int = 1024):
    """
    Upload any CSV file and automatically transform it to the standard 20-feature format.
    This prepares the data for live simulation.
    """
    global CUSTOM_DATA_STREAM, CUSTOM_CURRENT_INDEX, CUSTOM_TOTAL_RECORDS, CUSTOM_TRANSFORMATION_INFO
    
    try:
        # Read the uploaded CSV
        df = pd.read_csv(file.file)
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Uploaded file is empty")
        
        # Transform the data to standard 20-feature format
        transformer = DataTransformer(window_size=window_size)
        transformed_df, transform_info = transformer.transform_to_standard(df)
        
        # Store for simulation
        CUSTOM_DATA_STREAM = transformed_df.to_dict('records')
        CUSTOM_CURRENT_INDEX = 0
        CUSTOM_TOTAL_RECORDS = len(CUSTOM_DATA_STREAM)
        CUSTOM_TRANSFORMATION_INFO = transform_info
        
        print(f"📦 Custom data loaded: {CUSTOM_TOTAL_RECORDS} records")
        print(f"🔄 Transformation: {transform_info['detected_format']}")
        
        return {
            "success": True,
            "filename": file.filename,
            "original_rows": transform_info.get('original_rows', len(df)),
            "original_columns": transform_info.get('original_columns', df.shape[1]),
            "transformed_rows": CUSTOM_TOTAL_RECORDS,
            "transformed_columns": 20,
            "transformation_type": transform_info['detected_format'],
            "transformation_details": transform_info,
            "ready_for_simulation": True
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/simulate-custom/reset")
def reset_custom_simulation():
    """Reset the custom data simulation to the beginning"""
    global CUSTOM_CURRENT_INDEX
    CUSTOM_CURRENT_INDEX = 0
    
    return {
        "message": "Custom simulation reset",
        "total_records": CUSTOM_TOTAL_RECORDS,
        "transformation_info": CUSTOM_TRANSFORMATION_INFO
    }


@app.get("/simulate-custom/next")
def get_next_custom_reading(db: Session = Depends(get_db)):
    """Get the next row from custom uploaded data and run prediction"""
    global CUSTOM_CURRENT_INDEX
    
    if not CUSTOM_DATA_STREAM:
        raise HTTPException(status_code=404, detail="No custom data loaded. Upload a file first using /preprocess-upload")
    
    if CUSTOM_CURRENT_INDEX >= len(CUSTOM_DATA_STREAM):
        return {"finished": True, "total_records": CUSTOM_TOTAL_RECORDS}
    
    # Get the row
    row = CUSTOM_DATA_STREAM[CUSTOM_CURRENT_INDEX]
    CUSTOM_CURRENT_INDEX += 1
    
    # Extract features (20 values in standard format)
    feature_values = list(row.values())[:20]
    
    # Run prediction
    features_arr = np.array(feature_values).reshape(1, -1)
    features_scaled = scaler.transform(features_arr)
    recon = model.inverse_transform(model.transform(features_scaled))
    mse = np.mean(np.square(features_scaled - recon), axis=1)[0]
    
    thresh_val = float(threshold)
    mse_val = float(mse)
    status = "CRITICAL_FAILURE" if mse_val > thresh_val else "HEALTHY"
    
    # Save to DB
    new_record = sql_models.PredictionRecord(
        input_features=feature_values,
        anomaly_score=mse_val,
        threshold=thresh_val,
        status=status
    )
    db.add(new_record)
    db.commit()
    db.refresh(new_record)
    
    return {
        "finished": False,
        "index": CUSTOM_CURRENT_INDEX,
        "total_records": CUSTOM_TOTAL_RECORDS,
        "anomaly_score": mse_val,
        "threshold": thresh_val,
        "status": status,
        "alert": bool(mse_val > thresh_val)
    }


@app.get("/simulate-custom/status")
def get_custom_simulation_status():
    """Get the status of custom data simulation"""
    return {
        "data_loaded": len(CUSTOM_DATA_STREAM) > 0,
        "total_records": CUSTOM_TOTAL_RECORDS,
        "current_index": CUSTOM_CURRENT_INDEX,
        "progress_percent": round((CUSTOM_CURRENT_INDEX / CUSTOM_TOTAL_RECORDS * 100), 1) if CUSTOM_TOTAL_RECORDS > 0 else 0,
        "transformation_info": CUSTOM_TRANSFORMATION_INFO
    }


# ============================================================
# EXCEL EXPORT ENDPOINTS
# ============================================================
from fastapi.responses import StreamingResponse
from fastapi import Body
from io import BytesIO

@app.post("/export/batch-results")
async def export_batch_results_excel(results: dict = Body(...)):
    """Export batch analysis results to Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        print(f"📥 Export request received with {len(results.get('results', []))} results")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batch Analysis Results"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00C9FF", end_color="00C9FF", fill_type="solid")
        
        # Write headers
        headers = ["Row ID", "Anomaly Score", "Status", "Threshold"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        results_list = results.get("results", [])
        for row_idx, result in enumerate(results_list, 2):
            ws.cell(row=row_idx, column=1, value=result.get("index", row_idx - 2))
            ws.cell(row=row_idx, column=2, value=result.get("score", 0))
            status = "CRITICAL FAILURE" if result.get("is_anomaly", False) else "NORMAL"
            ws.cell(row=row_idx, column=3, value=status)
            ws.cell(row=row_idx, column=4, value=float(threshold))
            
            # Color code anomalies
            if result.get("is_anomaly", False):
                for c in range(1, 5):
                    ws.cell(row=row_idx, column=c).fill = PatternFill(
                        start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
                    )
        
        # Auto-adjust column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"batch_analysis_{results.get('filename', 'results').replace('.csv', '')}.xlsx"
        print(f"✅ Export file created: {filename}")
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        import traceback
        print(f"❌ Export error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@app.get("/export/simulation-history")
def export_simulation_history_excel(db: Session = Depends(get_db)):
    """Export all prediction history from database to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Get all records
        records = db.query(sql_models.PredictionRecord).order_by(
            sql_models.PredictionRecord.id.desc()
        ).limit(10000).all()
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Prediction History"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00C9FF", end_color="00C9FF", fill_type="solid")
        
        # Write headers
        headers = ["ID", "Anomaly Score", "Threshold", "Status", "Created At"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Write data
        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=record.id)
            ws.cell(row=row_idx, column=2, value=record.anomaly_score)
            ws.cell(row=row_idx, column=3, value=record.threshold)
            ws.cell(row=row_idx, column=4, value=record.status)
            ws.cell(row=row_idx, column=5, value=str(record.created_at) if record.created_at else "")
            
            # Color code anomalies
            if record.status == "CRITICAL_FAILURE":
                for col in range(1, 6):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"
                    )
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from datetime import datetime
        filename = f"prediction_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
